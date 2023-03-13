import openpyxl as xl
import copy
import random
import math
from django.conf import settings
import os

#########################################################################################

#FROM SANDESH

priority_list = [["The Dark Knight"], ["The Dark Knight"], ["The Dark Knight"]]
new_rated_movies = [
                    ["The Dark Knight", "2000", "Christopher Nolan", ["Action", "Crime", "Drama"], ["Christian Bale", "Heath Ledger", "Aaron Eckhart", "Michael Caine]"], 7]
                    ]


#########################################################################################

def priority(priority_list, new_rated_movies_only):

    #REQUIRED VARIABLES CALCULATIONS
    path_read = os.path.join(settings.BASE_DIR, "acounts/files/movie_list.xlsx")
    movie_list = []
    new_rated_movies = []
    level_of_rated = {}
    new_priority_list = [[], [], []]

  
    #name, 2 similar to 4 similar list
    for n in range(len(new_rated_movies_only)):
        level_of_rated.update({new_rated_movies[n]: [[], [], []]})


    #loading table from excel
    movies_excel = xl.load_workbook(path_read)
    movies_sheet = movies_excel["movie_list"]


    #converting table to movie_list
    for row in range(2, movies_sheet.max_row + 1):
        particular_movie = []
        genre = []
        cast = []
        for column in range(1, movies_sheet.max_column):
            cell = movies_sheet.cell(row, column)
            if column >= 3 and column <= 5 and cell.value != None:
                genre.append(cell.value)
            elif column >= 7 and column <= 10 and cell.value != None:
                cast.append(cell.value)
            elif cell.value != None or column == movies_sheet.max_column:
                particular_movie.append(cell.value)
        particular_movie.append(genre)
        particular_movie.append(cast)
        movie_list.append(particular_movie)


    #converting table to new_rated_movies categories
    for row in range(2, movies_sheet.max_row + 1):
        particular_movie = []
        genre = []
        cast = []
        for column in range(1, movies_sheet.max_column):
            cell = movies_sheet.cell(row, column)
            if cell.value in new_rated_movies_only:
                if column >= 3 and column <= 5 and cell.value != None:
                    genre.append(cell.value)
                elif column >= 7 and column <= 10 and cell.value != None:
                    cast.append(cell.value)
                elif cell.value != None or column == movies_sheet.max_column:
                    particular_movie.append(cell.value)
        particular_movie.append(genre)
        particular_movie.append(cast)
        new_rated_movies.append(particular_movie)


    #permutations
    perm = [[2, 3, 4, 1], [2, 4, 3, 1], [3, 2, 4, 1], [3, 4, 2, 1], [4, 2, 3, 1], [4, 3, 2, 1]]

    #########################################################################################

    #BUILDING TREE AND EXTRACTING DATA FROM KD - TREE

    #class for tree node
    class TreeNode:

        def __init__(self, data):
            self.data = data
            self.children = []
            self.parent = None

        def add_child(self, child):
            child.parent = self
            self.children.append(child)

        def get_level(self):
            level = 0
            p = self.parent
            while p:
                level += 1
                p = p.parent
            return level

        def print_tree(self):
            spaces = ' ' * self.get_level() * 3
            
            #root lai nothing, parent lai "|-->" , leaf lai "#"
            if self.parent:
                if not self.children:
                    prefix = spaces + "# "
                else:
                    prefix = spaces + "|--> "
            else:
                prefix = ""

            print(prefix + self.data)

            #recursion printing
            if self.children:
                for child in self.children:
                    child.print_tree()

        def Data(self):
            return self.data

        def childrenNode(self):
            return self.children
        
        def childrenNames(self):
            children_names = []
            for n in range(len(self.children)):
                children_names.append(self.children[n].data)
            return children_names


    #layer for fixed movie
    def layerFixed(a, layer, n, this_node, current_input):
        
        for x in range(len(movie_list)):

            #getting current depth of node
            layer = this_node.get_level() + 1
            
            #checking if all previous aspects matches
            count = 0
            for m in range(layer -1):
                #if have to compare list
                if type(movie_list[x][perm[n][m]]) == list:
                    common = list(set(movie_list[x][perm[n][m]]).intersection(movie_list[a][perm[n][m]]))
                    if len(common) != 0:
                        count += 1
                #if have to compare string
                elif movie_list[x][perm[n][m]] == movie_list[a][perm[n][m]]:
                    count += 1

            #if matches then start adding nodes
            if count == layer - 1:

                #if not leaf
                if layer < len(movie_list[0]):
                    new_node_name = copy.deepcopy(movie_list[x][perm[n][layer - 1]])
                
                #if leaf
                else:
                    new_node_name = copy.deepcopy(movie_list[x][0])

                children_names = this_node.childrenNames()
                children_nodes = this_node.childrenNode()

                #if already node exists or not
                exist = False
                y = 0
                while y < len(children_names):
                    
                    #if node exists just copy node
                    if new_node_name == children_names[y]:
                        exist = True
                        child_node = children_nodes[y]
                    y += 1

                #if no node exist, create new
                if not exist:
                    child_node = TreeNode(new_node_name)
                    this_node.add_child(child_node)

                #updating the names in level of rated
                if layer >= 3 and layer <= 5:
                    if movie_list[x][0] not in level_of_rated[current_input[0]][layer - 3]:
                        level_of_rated[current_input[0]][layer - 3].append(movie_list[x][0])
                
                #recursion until fixed leaf appears
                if layer < len(movie_list[0]) and child_node.Data() == current_input[perm[n][layer - 1]]:
                    layerFixed(x, layer, n, child_node, current_input)
        

    #building all trees of all permutations
    def fixed_build_tree():
        for p in range(len(new_rated_movies)):
            for m in range(len(movie_list)):
                if new_rated_movies[p][0] == movie_list[m][0]:
                    current_input = movie_list[m]

                    fixed_movie_tree = []

                    print("\n", movie_list[m][0])

                    for n in range(len(perm)):

                        print("perm", n + 1)

                        #creating trees one by one of different permutations
                        fixed_movie_tree.append(TreeNode("Movies"))
                        this_perm = fixed_movie_tree[n]
                        
                        a = 0 #just random value
                        layer = 1 #just random value
                    
                        layerFixed(a, layer, n, this_perm, current_input)


    #calling built function
    fixed_build_tree()

    #########################################################################################

    #COMBINING AND PRIORITY ACCORDING TO RATING AND SELECTION

    #combined list of all movies
    for new_movie in new_rated_movies:
        rate_value = new_movie[5]
        for list_index in range(3):
            for listed_movie in level_of_rated[new_movie[0]][list_index]:
                
                if ( (list_index == 2 or list_index == 1) and rate_value >= 7 ) or ( list_index == 0 and rate_value >= 5 ):
                    if listed_movie not in priority_list[list_index]:
                        new_priority_list[list_index].append(listed_movie)

    return new_priority_list


def recommended_movies(priority_list):
    
    #REQUIRED VARIABLES CALCULATIONS
    path_read = os.path.join(settings.BASE_DIR, "acounts/files/movie_list.xlsx")
    movie_list = []
    selected = []
    selected_movies = {}


    #loading table from excel
    movies_excel = xl.load_workbook(path_read)
    movies_sheet = movies_excel["movie_list"]


    #converting table to movie_list
    for row in range(2, movies_sheet.max_row + 1):
        particular_movie = []
        genre = []
        cast = []
        for column in range(1, movies_sheet.max_column):
            cell = movies_sheet.cell(row, column)
            if column >= 3 and column <= 5 and cell.value != None:
                genre.append(cell.value)
            elif column >= 7 and column <= 10 and cell.value != None:
                cast.append(cell.value)
            elif cell.value != None or column == movies_sheet.max_column:
                particular_movie.append(cell.value)
        particular_movie.append(genre)
        particular_movie.append(cast)
        movie_list.append(particular_movie)


    #ten random selection algorithm
    remaining_selection = 10

    #for 4 similar extraction
    if len(priority_list[2]) >= 6:
        while remaining_selection > 7:
            value = random.choice(priority_list[2])
            if value not in selected:
                selected.append(value)
                remaining_selection -= 1
                priority_list[2].remove(value)
                if value in priority_list[1]:
                    priority_list[1].remove(value)
                if value in priority_list[0]:
                    priority_list[0].remove(value)
    else:
        while remaining_selection > 10 - math.floor(len(priority_list[2])/2):
            value = random.choice(priority_list[2])
            if value not in selected:
                selected.append(value)
                remaining_selection -= 1
                priority_list[2].remove(value)
                if value in priority_list[1]:
                    priority_list[1].remove(value)
                if value in priority_list[0]:
                    priority_list[0].remove(value)
    print("\n4 Similar Selection Done: ", 10 - remaining_selection)

    #for 3 similar extraction
    extraction_no = remaining_selection - 2
    if len(priority_list[1]) >= extraction_no * 2:
        while remaining_selection > 2:
            value = random.choice(priority_list[1])
            if value not in selected:
                selected.append(value)
                remaining_selection -= 1
                priority_list[1].remove(value)
                if value in priority_list[0]:
                    priority_list[0].remove(value)
    else:
        temp_remaining_selection = copy.deepcopy(remaining_selection)
        while remaining_selection > temp_remaining_selection - math.floor(len(priority_list[1])/2):
            value = random.choice(priority_list[1])
            if value not in selected:
                selected.append(value)
                remaining_selection -= 1
                priority_list[1].remove(value)
                if value in priority_list[0]:
                    priority_list[0].remove(value)
    print("3 Similar Selection Done: ", extraction_no + 2 - remaining_selection)

    #for 2 similar extraction
    extraction_no = remaining_selection
    if len(priority_list[0]) >= extraction_no * 2:
        while remaining_selection > 0:
            value = random.choice(priority_list[0])
            if value not in selected:
                selected.append(value)
                remaining_selection -= 1
                priority_list[0].remove(value)
    else:
        temp_remaining_selection = copy.deepcopy(remaining_selection)
        while remaining_selection > temp_remaining_selection - math.floor(len(priority_list[0])/2):
            value = random.choice(priority_list[0])
            if value not in selected:
                selected.append(value)
                remaining_selection -= 1
                priority_list[0].remove(value)
    print("2 Similar Selection Done: ", extraction_no - remaining_selection)

    #for remaining random selection
    print("Random Selection Done   : ", remaining_selection, "\n")
    while len(selected) < 10:
        value = random.randrange(1, 999)
        if movie_list[value][0] not in selected:
            selected.append(movie_list[value][0])
        

    random.shuffle(selected)


    #making list of dictionaries of selected movies
    for selected_movie in selected:
        for movie in movie_list:
            if selected_movie == movie[0]:
                genre = []
                cast = []
                for gnere_index in range(len(movie[3])):
                    genre.append(movie[3][gnere_index])
                for cast_index in range(len(movie[4])):
                    cast.append(movie[4][cast_index])
                
                selected_movies.update({movie[0]: {"release_date": movie[1],
                                                "director": movie[2],
                                                "genre": genre,
                                                "cast": cast}
                                        })
                
    #PRINTING RECOMMENDED MOVIES
    for movies in selected:
        print(movies)


    return selected_movies

    #########################################################################################



    """
    selected_movies = {
                        {movie_name: {"release_date": sth,
                                                "director": sth,
                                                "genre": [genre_list],
                                                "cast": [cast_list]
                                    }
                        },

                        {movie_name: {"release_date": sth,
                                                "director": sth,
                                                "genre": [genre_list],
                                                "cast": [cast_list]
                                    }
                        }
                        }

    selected = [list_of_selected_movies_name]
    """

